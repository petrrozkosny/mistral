# ==============================================================================
# SKRIPT PRO GENEROVÁNÍ FORMULÁŘŮ CESTOVNÍCH NÁHRAD
# ==============================================================================
#
# Tento skript automaticky vytváří formuláře pro cestovní náhrady z Excel souboru
# se seznamem účastníků. Formuláře jsou rozděleny podle výkonnostních skupin.
#
# POUŽITÍ:
#   1. Připravte vstupní Excel soubor se sloupci:
#      - Příjmení, Jméno, Číslo OP, Bydliště, SPZ auta řidiče,
#      - Startuji ve výkonnostní skupině
#   2. Upravte KONFIGURACI níže (název akce, místo, termín, klub)
#   3. Spusťte skript: python app.py
#   4. Výstupní soubor se vytvoří ve stejné složce
#
# ==============================================================================

# Import potřebných knihoven
import pandas as pd  # Práce s Excel soubory a daty
from datetime import datetime  # Práce s datumy
import re  # Vyhledávání textu pomocí regulárních výrazů
from openpyxl import Workbook  # Vytváření a úprava Excel souborů


# ==============================================================================
# KONFIGURACE - UPRAVTE TYTO HODNOTY PODLE POTŘEBY
# ==============================================================================

# Názvy souborů
VSTUPNI_SOUBOR = "dummy_seznam.xlsx"  # Soubor se seznamem účastníků
VYSTUPNI_SOUBOR = "vystup_cestovni_nahrady.xlsx"  # Výstupní soubor s formuláři

# Informace o akci
NAZEV_KLUBU = "SK Moravia Brno"  # Název klubu/organizace
MISTO_KONANI = "Brno, 23. ročník turnaje stolního tenisu vozíčkářů"  # Místo konání
NAZEV_AKCE = '„Memoriál Vojtěcha Vašíčka"'  # Název akce
TERMIN = "2025-11-08"  # Datum akce (formát: YYYY-MM-DD)
SAZBA_KM = 2  # Náhrada za 1 km v Kč


# ==============================================================================
# POMOCNÉ FUNKCE
# ==============================================================================


def vytvor_formular(excel_list, zacatek_radku, osoba, datum_akce):
    """
    Vytvoří jeden formulář pro cestovní náhrady.

    Co dělá:
        - Vyplní formulář s osobními údaji (jméno, příjmení, adresa, OP)
        - Přidá vzorce pro automatický výpočet kilometrů a částky
        - Formulář má vždy stejnou strukturu (14 řádků)

    Parametry:
        excel_list     - list v Excelu, kam se formulář zapíše
        zacatek_radku  - od kterého řádku začít (aby se formuláře nevypsaly přes sebe)
        osoba          - data o osobě (příjmení, jméno, bydliště, OP, SPZ)
        datum_akce     - datum akce

    Vrací:
        Číslo řádku, kde formulář končí (abychom věděli, kde začít další formulář)
    """

    # Získání dat o osobě
    prijmeni = osoba.get("Příjmení", "")
    jmeno = osoba.get("Jméno", "")
    adresa = osoba.get("Bydliště", "")
    cislo_op = str(osoba.get("Číslo OP", ""))

    # Automatické nalezení PSČ v adrese
    # Hledá vzor: 3 číslice, volitelná mezera, 2-3 číslice
    # Příklad: "Brno 625 00" najde "62500"
    psc = ""
    if pd.notna(adresa):
        vysledek = re.search(r"(\d{3}\s?\d{2,3})", str(adresa))
        if vysledek:
            psc = vysledek.group(1).replace(" ", "")

    # Aktuální řádek, kam zapisujeme
    r = zacatek_radku

    # 2 prázdné řádky na začátek
    r += 2

    # Místo konání akce
    excel_list.cell(row=r, column=1).value = "Místo konání:"
    excel_list.cell(row=r, column=3).value = f" {MISTO_KONANI} "
    excel_list.cell(row=r, column=8).value = "……………………………………………"
    r += 1

    # Název akce a místo pro podpis
    excel_list.cell(row=r, column=4).value = f" {NAZEV_AKCE}"
    excel_list.cell(row=r, column=7).value = "Podpis odpovědného pracovníka"
    r += 1

    # Termín akce
    excel_list.cell(row=r, column=1).value = "Termín:"
    excel_list.cell(row=r, column=3).value = datum_akce
    r += 1

    # Prázdný řádek
    r += 1

    # Řádek s příjmením a hlavičkou tabulky cest
    excel_list.cell(row=r, column=1).value = "Příjmení:"
    excel_list.cell(row=r, column=2).value = prijmeni
    excel_list.cell(row=r, column=4).value = "Místo odjezdu - cíl cesty"
    excel_list.cell(row=r, column=6).value = "Počet km"
    excel_list.cell(row=r, column=7).value = "Celkem km"
    excel_list.cell(row=r, column=8).value = "SPZ"
    excel_list.cell(row=r, column=9).value = "Spolucestující:"
    r += 1

    # Vzorec: celkové kilometry = tam + zpět
    excel_list.cell(row=r, column=7).value = f"=F{r}+F{r+3}"
    r += 1

    # Jméno
    excel_list.cell(row=r, column=1).value = "Jméno:"
    excel_list.cell(row=r, column=2).value = jmeno
    r += 1

    # Bydliště a pokračování tabulky
    excel_list.cell(row=r, column=1).value = "Bydliště:"
    excel_list.cell(row=r, column=2).value = adresa
    excel_list.cell(row=r, column=4).value = "Konec cesty"
    excel_list.cell(row=r, column=6).value = "Počet km"
    excel_list.cell(row=r, column=8).value = datum_akce
    excel_list.cell(row=r, column=9).value = "……………………………………."
    r += 1

    # Vzorec: zpáteční cesta = stejné km jako tam
    excel_list.cell(row=r, column=6).value = f"=F{r-3}"
    r += 1

    # Místo pro datum a podpis
    excel_list.cell(row=r, column=8).value = "Datum"
    excel_list.cell(row=r, column=9).value = "Podpis účtovatele - příjemce"
    r += 1

    # PSČ, potvrzení a celková částka
    excel_list.cell(row=r, column=1).value = "PSČ:"
    excel_list.cell(row=r, column=2).value = psc
    excel_list.cell(row=r, column=4).value = (
        "Potvrzuji tímto, že jsem údaje uvedl správně a byl mi\n"
        "vyplacen příspěvek na cestovní výlohy os. automobilem:"
    )
    excel_list.cell(row=r, column=8).value = "Celkem:"
    excel_list.cell(row=r, column=9).value = f"=G{r-5}*{SAZBA_KM}"  # Vzorec: km * sazba
    excel_list.cell(row=r, column=10).value = "Kč"
    r += 1

    # Číslo OP a částka slovy
    excel_list.cell(row=r, column=1).value = "Č.OP:"
    excel_list.cell(row=r, column=2).value = cislo_op
    excel_list.cell(row=r, column=4).value = "Slovy:"
    r += 1

    # Vrátíme číslo řádku, kde skončil formulář
    return r


def vytvor_vystupni_soubor():
    """
    Hlavní funkce programu.

    Co dělá:
        1. Načte seznam osob ze vstupního Excel souboru
        2. Rozdělí osoby do skupin (A, B, C podle výkonnosti)
        3. Vytvoří nový Excel se samostatnými listy pro každou skupinu
        4. Do každého listu vygeneruje formuláře pro všechny osoby z této skupiny
        5. Uloží hotový Excel soubor
    """

    # Načtení vstupních dat
    print(f"Načítám data z {VSTUPNI_SOUBOR}...")
    vsechny_osoby = pd.read_excel(VSTUPNI_SOUBOR)
    print(f"Načteno {len(vsechny_osoby)} osob")

    # Převod textového data na datum pro Excel
    datum_akce = pd.to_datetime(TERMIN)

    # Rozdělení osob podle sloupce "Startuji ve výkonnostní skupině" (A, B, C, ...)
    skupiny = vsechny_osoby.groupby("Startuji ve výkonnostní skupině")

    # Vytvoření nového Excel sešitu
    excel = Workbook()
    excel.remove(excel.active)  # Odstraníme prázdný list, který se vytvoří automaticky

    # Pro každou skupinu (A, B, C...) vytvoříme samostatný list
    for nazev_skupiny in sorted(skupiny.groups.keys()):
        print(f"\nVytvářím list pro Skupinu {nazev_skupiny}...")

        # Získání všech osob z této skupiny
        osoby_ve_skupine = skupiny.get_group(nazev_skupiny)

        # Vytvoření nového listu s názvem např. "Skupina A"
        list_excel = excel.create_sheet(title=f"Skupina {nazev_skupiny}")

        # Hlavička na prvním řádku
        list_excel.cell(row=1, column=1).value = (
            f"Cestovní náhrady - příspěvek {NAZEV_KLUBU}"
        )

        # Formuláře začínáme psát od řádku 2
        aktualni_radek = 2

        # Projdeme všechny osoby ve skupině a vytvoříme pro každou formulář
        for _, osoba in osoby_ve_skupine.iterrows():
            print(f"  - {osoba['Jméno']} {osoba['Příjmení']}")

            # Převedeme data osoby na slovník
            data_osoby = osoba.to_dict()

            # Vytvoříme formulář a uložíme si, kde skončil
            aktualni_radek = vytvor_formular(
                list_excel, aktualni_radek, data_osoby, datum_akce
            )

    # Uložení Excel souboru
    print(f"\nUkládám do {VYSTUPNI_SOUBOR}...")
    excel.save(VYSTUPNI_SOUBOR)

    # Výpis souhrnu
    print(
        f"Hotovo! Vytvořeno {len(vsechny_osoby)} formulářů ve {len(skupiny)} skupinách"
    )
    print(f"Výstupní soubor: {VYSTUPNI_SOUBOR}")


# ==============================================================================
# SPUŠTĚNÍ PROGRAMU
# ==============================================================================
# Tento blok se spustí, pouze když spustíte tento soubor přímo
if __name__ == "__main__":
    vytvor_vystupni_soubor()
